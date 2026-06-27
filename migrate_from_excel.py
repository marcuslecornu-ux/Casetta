"""
migrate_from_excel.py
---------------------
Imports historical data from the Casetta workbook into casetta.db.
Run once from the casetta-app folder:

    python3 migrate_from_excel.py

Safe to re-run — uses INSERT OR IGNORE so duplicates are skipped.
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKBOOK   = os.path.join(SCRIPT_DIR, "..", "Casetta - Workbook 2026 - Relmagined.xlsm")
DB_PATH    = os.path.join(SCRIPT_DIR, "casetta.db")

if not os.path.exists(WORKBOOK):
    print(f"ERROR: Workbook not found at:\n  {WORKBOOK}")
    sys.exit(1)

if not os.path.exists(DB_PATH):
    print("ERROR: casetta.db not found. Start the app once first to create the database.")
    sys.exit(1)

print(f"Reading workbook…")
wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
c = conn.cursor()


def to_date(val):
    """Convert Excel date/datetime to ISO string, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None


def safe_float(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


# ── 1. DRINK SALES ─────────────────────────────────────────────────────────────
# Col indices (0-based): 1=DATE, 2=UID, 6=ROOM, 7=ITEM, 8=ID, 9=TYPE,
#   10=Bottle/Glass, 11=BOTTLE qty, 12=GLASS qty, 13=QUANTITY, 14=DISCOUNT flag,
#   17=BottleSell price, 18=GlassSell price, 19=TotalCost

print("\nImporting Drink Sales…")
ws = wb["Drinks - Sales"]
sales_ok = 0
sales_skip = 0
sell_prices = {}  # item_id -> (bottle_price, glass_price) — collected for stock update

for row in ws.iter_rows(min_row=3, values_only=True):
    date_val = row[1]
    if not date_val:
        continue

    sale_date  = to_date(date_val)
    if not sale_date:
        continue

    uid        = str(row[2]) if row[2] else f"HIST-{sale_date}-{sales_ok}"
    room       = str(row[6]) if row[6] else "Unknown"
    item_name  = str(row[7]) if row[7] else ""
    item_id    = str(row[8]) if row[8] else ""
    category   = str(row[9]) if row[9] else ""

    # Unit type: col 10 is "Bottle/Glass" flag; if None, use whichever qty is non-zero
    unit_flag  = str(row[10]).strip() if row[10] else ""
    bottle_qty = safe_int(row[11])
    glass_qty  = safe_int(row[12])
    quantity   = safe_int(row[13]) or (bottle_qty + glass_qty) or 1

    if unit_flag in ("Bottle", "BOTTLE"):
        unit_type = "Bottle"
    elif unit_flag in ("Glass", "GLASS"):
        unit_type = "Glass"
    else:
        unit_type = "Glass" if glass_qty > 0 and bottle_qty == 0 else "Bottle"

    bottle_sell = safe_float(row[17])
    glass_sell  = safe_float(row[18])

    # Collect best-known prices for stock table
    if item_id and (bottle_sell > 0 or glass_sell > 0):
        existing = sell_prices.get(item_id, (0.0, 0.0))
        sell_prices[item_id] = (
            max(existing[0], bottle_sell),
            max(existing[1], glass_sell)
        )

    unit_price = glass_sell if unit_type == "Glass" else bottle_sell

    # Discount: col 14 is a flag (1 = hosted/free in hosting rooms, else normal)
    discount_flag = safe_int(row[14])
    is_hosted  = 1 if (discount_flag == 1 and "Hosting" in room) else 0
    discount_pct = 100.0 if is_hosted else 0.0

    total_cost_col = safe_float(row[19])  # this is purchase cost, not sell revenue
    if is_hosted:
        total_sale = 0.0
    elif unit_price > 0:
        total_sale = unit_price * quantity * (1 - discount_pct / 100)
    else:
        # No price in workbook — record sale with 0 (price can be set in app later)
        total_sale = 0.0

    try:
        c.execute("""
            INSERT OR IGNORE INTO drink_sales
            (uid, date, room, item_id, item_name, category, quantity, unit_type,
             unit_price, discount_pct, is_hosted, total_sale, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (uid, sale_date, room, item_id, item_name, category,
              quantity, unit_type, unit_price, discount_pct, is_hosted,
              total_sale, "Imported from workbook"))
        sales_ok += 1
    except Exception as e:
        sales_skip += 1

print(f"  ✓ {sales_ok} sales imported, {sales_skip} skipped")


# ── 2. UPDATE STOCK PRICES from collected sell prices ──────────────────────────
print("\nUpdating stock item prices from sales history…")
prices_updated = 0
for item_id, (bottle_p, glass_p) in sell_prices.items():
    result = c.execute("""
        UPDATE stock_items
        SET selling_price_bottle = MAX(selling_price_bottle, ?),
            selling_price_glass  = MAX(selling_price_glass, ?)
        WHERE id = ? AND (selling_price_bottle = 0 OR selling_price_glass = 0)
    """, (bottle_p, glass_p, item_id))
    if result.rowcount:
        prices_updated += 1
print(f"  ✓ Prices updated for {prices_updated} stock items")


# ── 3. EXPENSES ────────────────────────────────────────────────────────────────
# Col indices (0-based): 1=date, 4=UID, 5=CATEGORY, 6=SUB_CATEGORY,
#   7=COMMENTS, 8=STATUS, 10=MONTH, 11=YEAR, 13=TOTAL

print("\nImporting Expenses…")
ws2 = wb["Expenses - ADD"]
exp_ok = 0
exp_skip = 0

for row in ws2.iter_rows(min_row=3, values_only=True):
    date_val = row[1]
    if not date_val:
        continue

    exp_date = to_date(date_val)
    if not exp_date:
        continue

    amount = safe_float(row[13])
    if amount == 0:
        continue

    uid      = str(row[4]) if row[4] else f"HISTEXP-{exp_date}-{exp_ok}"
    category = str(row[5]) if row[5] else "SUPPLIERS"
    sub_cat  = str(row[6]) if row[6] else ""
    comments = str(row[7]) if row[7] else ""
    status   = str(row[8]) if row[8] else "Paid"
    month    = str(row[10]) if row[10] else ""
    year     = safe_int(row[11]) or datetime.strptime(exp_date, "%Y-%m-%d").year

    # Normalise status
    if status not in ("Paid", "Forecast"):
        status = "Paid" if "paid" in status.lower() else "Forecast"

    try:
        c.execute("""
            INSERT OR IGNORE INTO expenses
            (uid, date, category, sub_category, comments, amount, status, month, year)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (uid, exp_date, category, sub_cat, comments, amount, status, month, year))
        exp_ok += 1
    except Exception as e:
        exp_skip += 1

print(f"  ✓ {exp_ok} expenses imported, {exp_skip} skipped")


# ── 4. BOOKINGS ────────────────────────────────────────────────────────────────
# Col indices (0-based): 1=ENTRY_DATE, 2=NAME, 3=ROOM, 4=ARRIVAL,
#   5=NIGHTS, 7=DEPARTURE, 10=SOURCE, 11=SOURCE_CODE, 12=CONFIRMED, 13=COST

print("\nImporting Bookings…")
ws3 = wb["Empire Bookings"]
book_ok = 0
book_skip = 0

for row in ws3.iter_rows(min_row=3, values_only=True):
    arrival_val = row[4]
    name = str(row[2]) if row[2] else ""
    if not arrival_val or not name or name in ("NAME", ""):
        continue

    entry_date = to_date(row[1]) or datetime.today().date().isoformat()
    arrival    = to_date(arrival_val)
    if not arrival:
        continue

    nights     = safe_int(row[5]) or 1
    departure  = to_date(row[7])
    if not departure:
        arr_dt    = datetime.strptime(arrival, "%Y-%m-%d")
        departure = (arr_dt + timedelta(days=nights)).date().isoformat()

    room       = str(row[3]) if row[3] else ""
    source     = str(row[10]) if row[10] else ""
    src_code   = str(row[11]) if row[11] else ""
    confirmed  = str(row[12]) if row[12] else "Yes"
    total_cost = safe_float(row[13])

    # Normalise confirmed
    conf_lower = confirmed.lower()
    if "yes" in conf_lower or "confirm" in conf_lower:
        confirmed = "Yes"
    elif "cxl" in conf_lower or "cancel" in conf_lower:
        confirmed = "CXL"
    else:
        confirmed = "Prov"

    try:
        c.execute("""
            INSERT INTO bookings
            (entry_date, guest_name, room, arrival, num_nights, departure,
             source, source_code, confirmed, rate_type, total_cost, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (entry_date, name, room, arrival, nights, departure,
              source, src_code, confirmed, "RACK", total_cost,
              "Imported from workbook"))
        book_ok += 1
    except Exception as e:
        book_skip += 1

print(f"  ✓ {book_ok} bookings imported, {book_skip} skipped")


# ── 5. STOCK LEVELS from Inventory 26 ─────────────────────────────────────────
# Open balance is col 3, use latest non-zero balance for current stock
print("\nUpdating stock levels from Inventory 26…")
ws4 = wb["Inventory 26"]
stock_updated = 0

for row in ws4.iter_rows(min_row=4, values_only=True):
    item_id = str(row[0]) if row[0] else ""
    if not item_id or item_id == "ID":
        continue

    # Find the most recent non-None balance value (cols 3,6,9,12,15,18,21...)
    # Balance cols: 6=Jan bal, 9=Feb bal, 12=Mar, 15=Apr, 18=May, 21=Jun
    last_balance = safe_int(row[3])  # open balance as fallback
    for col_idx in [6, 9, 12, 15, 18, 21]:
        if col_idx < len(row) and row[col_idx] is not None:
            last_balance = safe_int(row[col_idx])

    # Only update if item exists in our DB
    result = c.execute(
        "UPDATE stock_items SET current_stock=? WHERE id=?",
        (last_balance, item_id)
    )
    if result.rowcount:
        stock_updated += 1

print(f"  ✓ Stock levels updated for {stock_updated} items")


# ── Done ───────────────────────────────────────────────────────────────────────
conn.commit()
conn.close()

print("\n" + "─"*50)
print("Migration complete! Restart the app to see your data.")
print("─"*50)
